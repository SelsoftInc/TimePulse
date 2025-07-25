import pytesseract
import pandas as pd
import json
import openai
import pdfplumber
import docx
import os
import re
from PIL import Image
from io import BytesIO
from flask import Flask, request, render_template
from werkzeug.utils import secure_filename
from langchain.chains import LLMChain
from langchain_community.chat_models import ChatOpenAI
from langchain_community.chat_models import AzureChatOpenAI
from langchain.prompts import PromptTemplate
from langchain_community.callbacks import get_openai_callback
import logging
from dotenv import load_dotenv
import warnings
from docx import Document
from docx.opc.constants import RELATIONSHIP_TYPE as RT
from PIL import Image
import io
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage

# Ignore warnings
warnings.filterwarnings("ignore")

# Load environment variables
load_dotenv()

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize OpenAI LLM
#llm = ChatOpenAI(model="gpt-4o", openai_api_key=os.getenv("OPENAI_API_KEY"), verbose=True)

# Retrieve Azure OpenAI credentials from environment variables
azure_openai_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
azure_openai_api_key = os.getenv("AZURE_OPENAI_API_KEY")
azure_openai_deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME")
azure_openai_api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2023-06-01-preview")

if not azure_openai_endpoint or not azure_openai_api_key or not azure_openai_deployment_name:
    raise ValueError("Azure OpenAI endpoint, API key, and deployment name must be set in the .env file.")

print("yay")
# Initialize Azure OpenAI GPT model with LangChain
llm = AzureChatOpenAI(
    azure_deployment=azure_openai_deployment_name,
    api_version=azure_openai_api_version,
    temperature=0.7,
    max_tokens=None,
    max_retries=2
)

# Flask app setup
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = "uploads"
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Allowed file types
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "pdf", "docx", "xlsx"}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_image(image_path):
    image = Image.open(image_path)
    return pytesseract.image_to_string(image)

def extract_text_from_pdf(pdf_path):
    text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Extract text if available
            page_text = page.extract_text()
            if page_text:
                text.append(page_text)

            # OCR fallback for pages without text
            try:
                image = page.to_image(resolution=300)
                img_bytes = image.original.convert("RGB")
                ocr_text = pytesseract.image_to_string(img_bytes)
                if ocr_text.strip():
                    text.append(ocr_text)
            except Exception as e:
                logger.warning(f"OCR failed for a page in PDF: {e}")

    return "\n".join(text)

def extract_text_from_docx(docx_path):
    text = []
    images_ocr_text = []

    doc = Document(docx_path)

    # Extract all paragraph text
    for para in doc.paragraphs:
        text.append(para.text)

    # Extract images from document
    rels = doc.part._rels
    for rel in rels:
        rel = rels[rel]
        if rel.reltype == RT.IMAGE:
            img_data = rel.target_part.blob
            img = Image.open(io.BytesIO(img_data))
            ocr_result = pytesseract.image_to_string(img)
            if ocr_result.strip():
                images_ocr_text.append(ocr_result)

    # Combine plain text and OCR results
    combined_text = "\n".join(text + images_ocr_text)
    return combined_text

def extract_text_from_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_text = []
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_text = [f"--- Sheet: {sheet} ---"]
        
        # Extract cell text
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            sheet_text.append("\t".join(row_values))
        
        # OCR any embedded images (if any)
        for image in ws._images:  # This is a private attribute but works
            try:
                if isinstance(image, ExcelImage):
                    img_bytes = image.ref.blob  # fallback if using openpyxl with patched blob
                else:
                    img_bytes = image._data()
                pil_img = Image.open(BytesIO(img_bytes))
                ocr_text = pytesseract.image_to_string(pil_img)
                if ocr_text.strip():
                    sheet_text.append(f"\n[Image OCR Text]\n{ocr_text}")
            except Exception as e:
                logger.warning(f"Failed to extract image or OCR from sheet {sheet}: {e}")
        
        all_text.append("\n".join(sheet_text))

    return "\n\n".join(all_text)

def process_file(file_path):
    ext = file_path.rsplit('.', 1)[1].lower()
    if ext in ["png", "jpg", "jpeg"]:
        return extract_text_from_image(file_path)
    elif ext == "pdf":
        return extract_text_from_pdf(file_path)
    elif ext == "docx":
        return extract_text_from_docx(file_path)
    elif ext == "xlsx":
        return extract_text_from_xlsx(file_path)
    else:
        raise ValueError("Unsupported file format")

def process_with_llm(text):
    prompt_template = PromptTemplate(
        input_variables=["text"],
        template="""
        Extract the following details from the timesheet provided:
        - Vendor Name (Infer from document headers or context)
        - Start Date
        - End Date
        - Total Hours (Billable, Non-Billable, Holiday Hours)
        - Invoice Amount
        Ensure the response is formatted as **valid JSON**.
        Timesheet Data:
        {text}
        
        JSON format is below
        {{
    "Vendor Name": ,
    "Start Date": ,
    "End Date": ",
    "Total Hours": [
        "Billable Project Hrs": ,
        "Non-Billable Project Hrs": ,
        "Time off/Holiday Hrs": 
    ],
    "Invoice Amount": 
}}
        """
    )

    chain = LLMChain(llm=llm, prompt=prompt_template)
    with get_openai_callback() as cb:
        feedback = chain.invoke({"text": text}, return_full_output=True)
    
    feedback_text = feedback.get("text", "")
    print("feedback",feedback_text)
    json_match = re.search(r"\{.*\}", feedback_text, re.DOTALL)
    json_string = json_match.group(0) if json_match else "{}"
    try:
        results = json.loads(json_string)
    except json.JSONDecodeError:
        logger.error("Invalid JSON extracted from GPT response.")
        results = {}
    return results

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('index.html', error="No file uploaded")
        file = request.files['file']
        if file.filename == '':
            return render_template('index.html', error="No selected file")
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            raw_text = process_file(file_path)
            print("raw_text",raw_text)
            structured_data = process_with_llm(raw_text)
            print("structured_data",structured_data)
            return render_template('index.html', data=structured_data)
    return render_template('index.html', data=None)

if __name__ == "__main__":
    app.run(debug=True)
