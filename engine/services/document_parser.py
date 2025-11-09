from langchain_community.document_loaders import (
    PyPDFLoader,
    UnstructuredImageLoader,
    UnstructuredWordDocumentLoader,
    CSVLoader
)
try:
    from langchain.text_splitter import RecursiveCharacterTextSplitter
except Exception:
    # Fallback: provide a minimal splitter if langchain is not installed.
    class RecursiveCharacterTextSplitter:
        def __init__(self, chunk_size=4000, chunk_overlap=200, length_function=len):
            self.chunk_size = int(chunk_size)
            self.chunk_overlap = int(chunk_overlap)
            self.length_function = length_function

        def split_text(self, text: str):
            """A very small fallback splitter that splits by characters with overlap.

            This is intentionally simple and only used when LangChain isn't available.
            """
            if not text:
                return []
            chunks = []
            step = max(1, self.chunk_size - self.chunk_overlap)
            for i in range(0, len(text), step):
                chunk = text[i:i + self.chunk_size]
                chunks.append(chunk)
                if i + self.chunk_size >= len(text):
                    break
            return chunks
from typing import List, Dict
import os
import tempfile
from pathlib import Path
from loguru import logger
import pandas as pd
from PIL import Image
import numpy as np
import cv2
import boto3
from botocore.exceptions import ClientError
import json
from config import get_settings
import io
try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

# Document conversion imports
try:
    from docx import Document as DocxDocument
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    import xlrd
except Exception:
    logger.warning("Document conversion libraries not available. DOC/XLS conversion will be skipped.")


class DocumentParser:
    """Handle parsing of different document types"""
    
    def __init__(self):
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=4000,
            chunk_overlap=200,
            length_function=len,
        )
        # Initialize Bedrock client (boto3). If AWS creds are provided in env, boto3 will use them.
        self.settings = get_settings()
        self.bedrock_runtime = None
        try:
            # If explicit AWS credentials provided in settings, pass them; otherwise rely on default chain
            if self.settings.AWS_ACCESS_KEY_ID and self.settings.AWS_SECRET_ACCESS_KEY:
                self.bedrock_runtime = boto3.client(
                    service_name='bedrock-runtime',
                    region_name=self.settings.AWS_REGION,
                    aws_access_key_id=self.settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=self.settings.AWS_SECRET_ACCESS_KEY,
                )
            else:
                self.bedrock_runtime = boto3.client(
                    service_name='bedrock-runtime',
                    region_name=self.settings.AWS_REGION,
                )
            logger.info("Initialized Bedrock runtime client for IDP")
        except Exception as e:
            logger.warning(f"Failed to initialize Bedrock client: {e}")
            self.bedrock_runtime = None
    
    async def parse_document(self, file_path: str, file_extension: str) -> str:
        """
        Parse document based on file type and extract text with preprocessing
        
        Args:
            file_path: Path to the document
            file_extension: File extension (pdf, png, docx, etc.)
        
        Returns:
            Extracted text content
        """
        try:
            logger.info(f"Parsing document: {file_path} with extension: {file_extension}")
            
            # Preprocess files based on type
            processed_file_path = await self._preprocess_document(file_path, file_extension)
            
            # Analyze the processed document (now always an upscaled PNG)
            text = await self.analyze_document(processed_file_path)
            
            # Clean up temporary processed file if different from original
            if processed_file_path != file_path:
                try:
                    os.remove(processed_file_path)
                    logger.info(f"Cleaned up temporary file: {processed_file_path}")
                except Exception as e:
                    logger.warning(f"Could not clean up temporary file {processed_file_path}: {e}")
            
            logger.info(f"Successfully extracted {len(text)} characters from document")
            return text
            
        except Exception as e:
            logger.error(f"Error parsing document: {str(e)}")
            raise
    
    async def _preprocess_document(self, file_path: str, file_extension: str) -> str:
        """
        Preprocess document based on type, converting to upscaled PNG for optimal OCR
        
        Args:
            file_path: Path to the document
            file_extension: File extension
            
        Returns:
            Path to processed PNG file
        """
        try:
            # Document formats that need conversion to PDF first
            if file_extension in ['doc', 'docx']:
                logger.info("Converting Word document to PDF...")
                pdf_path = await self._convert_word_to_pdf(file_path)
                png_path = await self._convert_pdf_to_png(pdf_path, upscale=self.settings.ENABLE_IMAGE_UPSCALING, dpi=self.settings.PDF_TO_PNG_DPI)
                os.remove(pdf_path)  # Clean up intermediate PDF
                return png_path
                
            elif file_extension in ['xls', 'xlsx']:
                logger.info("Converting Excel spreadsheet to PDF...")
                pdf_path = await self._convert_excel_to_pdf(file_path)
                png_path = await self._convert_pdf_to_png(pdf_path, upscale=self.settings.ENABLE_IMAGE_UPSCALING, dpi=self.settings.PDF_TO_PNG_DPI)
                os.remove(pdf_path)  # Clean up intermediate PDF
                return png_path
                
            elif file_extension == 'pdf':
                logger.info("Converting PDF to upscaled PNG...")
                return await self._convert_pdf_to_png(file_path, upscale=self.settings.ENABLE_IMAGE_UPSCALING, dpi=self.settings.PDF_TO_PNG_DPI)
                
            elif file_extension in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
                logger.info("Upscaling image...")
                return await self._upscale_image(file_path, method=self.settings.UPSCALING_METHOD, scale_factor=self.settings.UPSCALING_SCALE_FACTOR)
                
            else:
                # For text-based formats, convert to PDF then PNG
                logger.info(f"Converting {file_extension} document to PDF...")
                pdf_path = await self._convert_text_to_pdf(file_path, file_extension)
                png_path = await self._convert_pdf_to_png(pdf_path, upscale=self.settings.ENABLE_IMAGE_UPSCALING, dpi=self.settings.PDF_TO_PNG_DPI)
                os.remove(pdf_path)  # Clean up intermediate PDF
                return png_path
                
        except Exception as e:
            logger.error(f"Error preprocessing document: {e}")
            # Fallback: return original file if preprocessing fails
            logger.warning("Preprocessing failed, using original file")
            return file_path
    
    async def _convert_word_to_pdf(self, file_path: str) -> str:
        """Convert Word document to PDF"""
        try:
            from docx import Document as DocxDocument
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            
            # Create temporary PDF path
            pdf_path = file_path.rsplit('.', 1)[0] + '_converted.pdf'
            
            # Read Word document
            doc = DocxDocument(file_path)
            
            # Create PDF
            doc_pdf = SimpleDocTemplate(pdf_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            for para in doc.paragraphs:
                if para.text.strip():
                    story.append(Paragraph(para.text, styles['Normal']))
                    story.append(Spacer(1, 12))
            
            doc_pdf.build(story)
            logger.info(f"Converted Word to PDF: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            logger.error(f"Error converting Word to PDF: {e}")
            raise
    
    async def _convert_excel_to_pdf(self, file_path: str) -> str:
        """Convert Excel spreadsheet to PDF"""
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib import colors
            import openpyxl
            
            # Create temporary PDF path
            pdf_path = file_path.rsplit('.', 1)[0] + '_converted.pdf'
            
            # Read Excel file
            if file_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path)
            else:
                # For .xls files, we'd need xlrd, but let's use pandas as fallback
                import pandas as pd
                df = pd.read_excel(file_path)
                wb = None
            
            # Create PDF
            doc = SimpleDocTemplate(pdf_path, pagesize=letter)
            elements = []
            
            if wb:
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    data = []
                    
                    # Get all rows
                    for row in sheet.iter_rows(values_only=True):
                        data.append([str(cell) if cell is not None else '' for cell in row])
                    
                    if data:
                        table = Table(data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 14),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        elements.append(table)
            else:
                # Fallback for .xls files
                table = Table([df.columns.tolist()] + df.values.tolist())
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
            
            doc.build(elements)
            logger.info(f"Converted Excel to PDF: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            logger.error(f"Error converting Excel to PDF: {e}")
            raise
    
    async def _convert_text_to_pdf(self, file_path: str, file_extension: str) -> str:
        """Convert text-based documents to PDF"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            
            # Create temporary PDF path
            pdf_path = file_path.rsplit('.', 1)[0] + '_converted.pdf'
            
            # Read text content
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Create PDF
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            
            # Split content into lines that fit the page
            lines = content.split('\n')
            y_position = height - 50
            
            for line in lines:
                if y_position < 50:  # New page if needed
                    c.showPage()
                    y_position = height - 50
                
                c.drawString(50, y_position, line[:80])  # Limit line length
                y_position -= 15
            
            c.save()
            logger.info(f"Converted {file_extension} to PDF: {pdf_path}")
            return pdf_path
            
        except Exception as e:
            logger.error(f"Error converting {file_extension} to PDF: {e}")
            raise
    
    async def _convert_pdf_to_png(self, pdf_path: str, upscale: bool = True, dpi: int = 300) -> str:
        """Convert PDF to high-resolution PNG"""
        try:
            from pdf2image import convert_from_path
            
            # Create temporary PNG path
            png_path = pdf_path.rsplit('.', 1)[0] + '_highres.png'
            
            # Convert PDF to images
            images = convert_from_path(pdf_path, dpi=dpi, fmt='png')
            
            if not images:
                raise ValueError("No images extracted from PDF")
            
            # Use first page (or combine multiple pages if needed)
            img = images[0]
            
            # Upscale if requested
            if upscale:
                img = await self._upscale_image_from_pil(img)
            
            # Save the image
            img.save(png_path, 'PNG')
            logger.info(f"Converted PDF to PNG: {png_path} (DPI: {dpi})")
            return png_path
            
        except Exception as e:
            logger.error(f"Error converting PDF to PNG: {e}")
            raise
    
    async def _upscale_image(self, image_path: str, method: str = 'lanczos', scale_factor: float = 2.0) -> str:
        """Upscale an image using various methods"""
        try:
            # Read image
            img = cv2.imread(image_path)
            if img is None:
                # Fallback to PIL
                pil_img = Image.open(image_path)
                return await self._upscale_image_from_pil(pil_img, method, scale_factor, save_path=image_path)
            
            height, width = img.shape[:2]
            new_width = int(width * scale_factor)
            new_height = int(height * scale_factor)
            
            # Choose interpolation method
            if method == 'lanczos':
                interpolation = cv2.INTER_LANCZOS4
            elif method == 'cubic':
                interpolation = cv2.INTER_CUBIC
            elif method == 'linear':
                interpolation = cv2.INTER_LINEAR
            else:
                interpolation = cv2.INTER_LANCZOS4
            
            # Upscale
            upscaled = cv2.resize(img, (new_width, new_height), interpolation=interpolation)
            
            # Create output path
            output_path = image_path.rsplit('.', 1)[0] + '_upscaled.png'
            
            # Save
            cv2.imwrite(output_path, upscaled)
            logger.info(f"Upscaled image: {image_path} -> {output_path} ({width}x{height} -> {new_width}x{new_height})")
            return output_path
            
        except Exception as e:
            logger.error(f"Error upscaling image: {e}")
            # Return original if upscaling fails
            return image_path
    
    async def _upscale_image_from_pil(self, pil_img: Image.Image, method: str = 'lanczos', 
                                    scale_factor: float = 2.0, save_path: str = None) -> str:
        """Upscale image using PIL"""
        try:
            width, height = pil_img.size
            new_width = int(width * scale_factor)
            new_height = int(height * scale_factor)
            
            # Choose resampling filter
            if method == 'lanczos':
                resample = Image.Resampling.LANCZOS
            elif method == 'bicubic':
                resample = Image.Resampling.BICUBIC
            elif method == 'bilinear':
                resample = Image.Resampling.BILINEAR
            else:
                resample = Image.Resampling.LANCZOS
            
            # Upscale
            upscaled = pil_img.resize((new_width, new_height), resample)
            
            # Create output path
            if save_path:
                output_path = save_path.rsplit('.', 1)[0] + '_upscaled.png'
            else:
                # Create temporary file
                temp_fd, output_path = tempfile.mkstemp(suffix='_upscaled.png')
                os.close(temp_fd)
            
            # Save
            upscaled.save(output_path, 'PNG')
            logger.info(f"Upscaled image with PIL: {width}x{height} -> {new_width}x{new_height}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error upscaling image with PIL: {e}")
            raise

    async def _parse_image(self, file_path: str) -> str:
        """Deprecated: image parsing should use `analyze_document`. Kept for compatibility."""
        return await self.analyze_document(file_path)

    async def analyze_document(self, file_path: str, user_prompt: str | None = None) -> str:
        """Send a document (image/pdf/doc/spreadsheet) to Bedrock Claude via boto3 converse.

        Returns the textual analysis returned by the model.
        """
        if not self.bedrock_runtime:
            raise RuntimeError("Bedrock runtime client not initialized. Check AWS credentials and settings.")

        # Default prompt if none provided
        if not user_prompt:
            user_prompt = "Please analyze this document and extract all important information. Provide a structured response including type, key fields, tables, and a short summary."

        try:
            file_content, doc_format, is_image = self._encode_file(file_path)

            if is_image:
                content_block = {
                    "image": {
                        "format": doc_format,
                        "source": {"bytes": file_content}
                    }
                }
            else:
                content_block = {
                    "document": {
                        "format": doc_format,
                        "name": f"document-{doc_format}",
                        "source": {"bytes": file_content}
                    }
                }

            message = {
                "role": "user",
                "content": [
                    content_block,
                    {"text": user_prompt}
                ]
            }

            # Determine model id
            model_id = self.settings.CLAUDE_MODEL_ID or self.settings.BEDROCK_CLAUDE_MODEL or self.settings.LLM_MODEL_ID
            if not model_id:
                raise RuntimeError("No model id configured (CLAUDE_MODEL_ID, BEDROCK_CLAUDE_MODEL or LLM_MODEL_ID)")

            logger.info(f"Sending document to Bedrock model {model_id}")
            response = self.bedrock_runtime.converse(
                modelId=model_id,
                messages=[message],
                inferenceConfig={
                    "maxTokens": 4096,
                    "temperature": 0.1,
                    "topP": 0.9
                }
            )

            # Attempt to extract the textual response
            result = None
            try:
                result = response['output']['message']['content'][0]['text']
            except Exception:
                # Fallback: stringify the response
                result = json.dumps(response)

            logger.info(f"Bedrock returned {len(result)} characters")
            return result

        except FileNotFoundError:
            logger.error(f"File not found: {file_path}")
            raise
        except ClientError as e:
            logger.error(f"Bedrock API error: {e}")
            raise
        except Exception as e:
            logger.error(f"Error analyzing document: {e}")
            raise

    def _encode_file(self, file_path: str):
        """Read file and return bytes, format and whether it's an image.

        Supports: pdf, csv, doc, docx, xls, xlsx, html, txt, md, png, jpg, jpeg, gif, webp
        """
        file_ext = Path(file_path).suffix.lower()

        format_mapping = {
            '.pdf': 'pdf',
            '.csv': 'csv',
            '.doc': 'doc',
            '.docx': 'docx',
            '.xls': 'xls',
            '.xlsx': 'xlsx',
            '.html': 'html',
            '.htm': 'html',
            '.txt': 'txt',
            '.md': 'md',
            '.png': 'png',
            '.jpg': 'jpeg',
            '.jpeg': 'jpeg',
            '.gif': 'gif',
            '.webp': 'webp'
        }

        doc_format = format_mapping.get(file_ext)
        if not doc_format:
            raise ValueError(f"Unsupported file type: {file_ext}. Supported: pdf, csv, doc, docx, xls, xlsx, html, htm, txt, md, png, jpg, jpeg, gif, webp")

        with open(file_path, 'rb') as f:
            content = f.read()

        is_image = file_ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp']
        return content, doc_format, is_image
    
    async def _parse_pdf(self, file_path: str) -> str:
        """Parse PDF document"""
        # Prefer Bedrock IDP for PDFs
        return await self.analyze_document(file_path)
    
    async def _parse_docx(self, file_path: str) -> str:
        """Parse DOCX document"""
        return await self.analyze_document(file_path)
    
    async def _parse_spreadsheet(self, file_path: str, file_extension: str) -> str:
        """Parse CSV or XLSX spreadsheet"""
        return await self.analyze_document(file_path)
    
    def chunk_text(self, text: str) -> List[str]:
        """Split text into chunks for processing"""
        chunks = self.text_splitter.split_text(text)
        return chunks

    async def analyze_image_bytes(self, image_bytes: bytes, image_format: str = "png", user_prompt: str | None = None) -> str:
        if not self.bedrock_runtime:
            raise RuntimeError("Bedrock runtime client not initialized. Check AWS credentials and settings.")
        if not user_prompt:
            user_prompt = "Please analyze this document and extract all important information. Provide a structured response including type, key fields, tables, and a short summary."
        try:
            content_block = {
                "image": {
                    "format": image_format,
                    "source": {"bytes": image_bytes}
                }
            }
            message = {
                "role": "user",
                "content": [
                    content_block,
                    {"text": user_prompt}
                ]
            }
            model_id = self.settings.CLAUDE_MODEL_ID or self.settings.BEDROCK_CLAUDE_MODEL or self.settings.LLM_MODEL_ID
            if not model_id:
                raise RuntimeError("No model id configured (CLAUDE_MODEL_ID, BEDROCK_CLAUDE_MODEL or LLM_MODEL_ID)")
            response = self.bedrock_runtime.converse(
                modelId=model_id,
                messages=[message],
                inferenceConfig={
                    "maxTokens": 4096,
                    "temperature": 0.1,
                    "topP": 0.9
                }
            )
            result = None
            try:
                result = response['output']['message']['content'][0]['text']
            except Exception:
                result = json.dumps(response)
            return result
        except Exception as e:
            logger.error(f"Error analyzing image bytes: {e}")
            raise

    def _pdf_to_png_pages(self, file_path: str, dpi: int = 180) -> List[bytes]:
        if not fitz:
            raise RuntimeError("PyMuPDF (fitz) not installed; cannot render PDF to PNG.")
        try:
            doc = fitz.open(file_path)
            images: List[bytes] = []
            for page_index in range(len(doc)):
                page = doc.load_page(page_index)
                zoom = dpi / 72.0
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                img_bytes = pix.tobytes("png")
                images.append(img_bytes)
            doc.close()
            logger.info(f"Rendered {len(images)} PNG page(s) from PDF")
            return images
        except Exception as e:
            logger.error(f"Failed to render PDF to PNG: {e}")
            raise

    async def analyze_document_pages_as_png(self, file_path: str, user_prompt: str | None = None) -> str:
        ext = Path(file_path).suffix.lower().lstrip('.')
        if ext != 'pdf':
            return await self.analyze_document(file_path, user_prompt)
        pages = self._pdf_to_png_pages(file_path)
        if not pages:
            return ""
        aggregated = []
        for idx, img in enumerate(pages, start=1):
            try:
                text = await self.analyze_image_bytes(img, "png", user_prompt)
                aggregated.append(text)
            except Exception as e:
                logger.warning(f"Skipping page {idx} due to error: {e}")
        return "\n\n".join(aggregated).strip()
